import csv
import sqlite3
from datetime import date
from pathlib import Path
from collections import Counter, defaultdict

from openpyxl import Workbook, load_workbook

from modules.finance.repositories.credit_card_expense_repository import (
    CreditCardExpenseRepository,
)
from modules.finance.repositories.credit_card_invoice_repository import (
    CreditCardInvoiceRepository,
)
from modules.finance.services.credit_card_invoice_service import (
    CreditCardInvoiceService,
)


EXPORT_COLUMNS = [
    "category_id",
    "original_description",
    "effective_description",
    "original_purchase_date",
    "effective_purchase_date",
    "original_amount_cents",
    "effective_amount_cents",
    "installment_number",
    "installment_total",
    "installment_group_id",
    "source_type",
    "source_reference",
    "notes",
]


class CreditCardPortableDataService:
    def __init__(
            self,
            username: str,
    ) -> None:
        self.expense_repository = CreditCardExpenseRepository(username)
        self.invoice_repository = CreditCardInvoiceRepository(username)
        self.invoice_service = CreditCardInvoiceService()

    def exportar_csv(
            self,
            credit_card: dict,
            destino: str,
    ) -> None:
        lancamentos = self.expense_repository.listar_por_cartao(
            credit_card["id"]
        )

        with open(destino, "w", newline="", encoding="utf-8-sig") as arquivo:
            writer = csv.DictWriter(
                arquivo,
                fieldnames=EXPORT_COLUMNS,
                delimiter=";",
            )

            writer.writeheader()

            for lancamento in lancamentos:
                writer.writerow(
                    {
                        coluna: lancamento.get(coluna)
                        for coluna in EXPORT_COLUMNS
                    }
                )

    def exportar_excel(
            self,
            credit_card: dict,
            destino: str,
    ) -> None:
        lancamentos = self.expense_repository.listar_por_cartao(
            credit_card["id"]
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "lancamentos_cartao"

        worksheet.append(EXPORT_COLUMNS)

        for lancamento in lancamentos:
            worksheet.append(
                [
                    lancamento.get(coluna)
                    for coluna in EXPORT_COLUMNS
                ]
            )

        workbook.save(destino)

    def exportar_db(
            self,
            credit_card: dict,
            destino: str,
    ) -> None:
        destino_path = Path(destino)

        if destino_path.exists():
            destino_path.unlink()

        conexao = sqlite3.connect(destino)
        cursor = conexao.cursor()

        cursor.execute(
            """
            CREATE TABLE exported_credit_card_expenses (
                category_id INTEGER,
                original_description TEXT,
                effective_description TEXT,
                original_purchase_date TEXT,
                effective_purchase_date TEXT,
                original_amount_cents INTEGER,
                effective_amount_cents INTEGER,
                installment_number INTEGER,
                installment_total INTEGER,
                installment_group_id TEXT,
                source_type TEXT,
                source_reference TEXT,
                notes TEXT
            )
            """
        )

        lancamentos = self.expense_repository.listar_por_cartao(
            credit_card["id"]
        )

        for lancamento in lancamentos:
            cursor.execute(
                """
                INSERT INTO exported_credit_card_expenses (
                    category_id,
                    original_description,
                    effective_description,
                    original_purchase_date,
                    effective_purchase_date,
                    original_amount_cents,
                    effective_amount_cents,
                    installment_number,
                    installment_total,
                    installment_group_id,
                    source_type,
                    source_reference,
                    notes
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                tuple(
                    lancamento.get(coluna)
                    for coluna in EXPORT_COLUMNS
                ),
            )

        conexao.commit()
        conexao.close()

    def importar_csv(
            self,
            credit_card: dict,
            origem: str,
            category_id: int = 1,
    ) -> int:
        with open(origem, "r", newline="", encoding="utf-8-sig") as arquivo:
            reader = csv.DictReader(
                arquivo,
                delimiter=";",
            )

            rows = [
                dict(row)
                for row in reader
            ]

        return self._importar_rows(
            credit_card=credit_card,
            rows=rows,
            category_id=category_id,
        )

    def importar_excel(
            self,
            credit_card: dict,
            origem: str,
            category_id: int = 1,
    ) -> int:
        workbook = load_workbook(origem)
        worksheet = workbook.active

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        rows = []

        for linha in worksheet.iter_rows(
                min_row=2,
                values_only=True,
        ):
            rows.append(
                dict(zip(headers, linha))
            )

        return self._importar_rows(
            credit_card=credit_card,
            rows=rows,
            category_id=category_id,
        )

    def importar_db(
            self,
            credit_card: dict,
            origem: str,
            category_id: int = 1,
    ) -> int:
        conexao = sqlite3.connect(origem)
        conexao.row_factory = sqlite3.Row

        rows = conexao.execute(
            """
            SELECT *
            FROM exported_credit_card_expenses
            """
        ).fetchall()

        conexao.close()

        return self._importar_rows(
            credit_card=credit_card,
            rows=[dict(row) for row in rows],
            category_id=category_id,
        )

    def _gerar_assinatura_row(
            self,
            row: dict,
    ) -> tuple:
        original_description = row.get("original_description") or row.get("effective_description")
        original_purchase_date = row.get("original_purchase_date") or row.get("effective_purchase_date")
        original_amount_cents = int(row.get("original_amount_cents") or row.get("effective_amount_cents") or 0)
        installment_number = int(row.get("installment_number") or 1)
        installment_total = int(row.get("installment_total") or 1)

        return (
            original_description,
            original_purchase_date,
            original_amount_cents,
            installment_number,
            installment_total,
        )

    def _importar_rows(
            self,
            credit_card: dict,
            rows: list[dict],
            category_id: int,
    ) -> int:
        total_importado = 0

        total_por_assinatura = Counter(
            self._gerar_assinatura_row(row)
            for row in rows
        )

        ja_importados_agora = defaultdict(int)

        for row in rows:
            assinatura = self._gerar_assinatura_row(row)

            total_no_banco = self.expense_repository.contar_lancamentos_por_assinatura(
                credit_card_id=credit_card["id"],
                original_description=assinatura[0],
                original_purchase_date=assinatura[1],
                original_amount_cents=assinatura[2],
                installment_number=assinatura[3],
                installment_total=assinatura[4],
            )

            if total_no_banco + ja_importados_agora[assinatura] >= total_por_assinatura[assinatura]:
                continue

            original_description = row.get("original_description") or row.get("effective_description")
            original_purchase_date = row.get("original_purchase_date") or row.get("effective_purchase_date")
            original_amount_cents = int(row.get("original_amount_cents") or row.get("effective_amount_cents") or 0)

            effective_description = row.get("effective_description") or original_description
            effective_purchase_date = row.get("effective_purchase_date") or original_purchase_date
            effective_amount_cents = int(row.get("effective_amount_cents") or original_amount_cents)

            installment_number = int(row.get("installment_number") or 1)
            installment_total = int(row.get("installment_total") or 1)

            purchase_date = date.fromisoformat(effective_purchase_date)

            invoice_year, invoice_month = self.invoice_service.calcular_mes_fatura(
                purchase_date=purchase_date,
                closing_day=credit_card["closing_day"],
            )

            closing_date = self.invoice_service.montar_data_segura(
                invoice_year,
                invoice_month,
                credit_card["closing_day"],
            )

            due_date = self.invoice_service.montar_data_segura(
                invoice_year,
                invoice_month,
                credit_card["due_day"],
            )

            invoice = self.invoice_repository.buscar_por_cartao_mes(
                credit_card_id=credit_card["id"],
                invoice_year=invoice_year,
                invoice_month=invoice_month,
            )

            if invoice is None:
                invoice_id = self.invoice_repository.criar_fatura(
                    credit_card_id=credit_card["id"],
                    invoice_year=invoice_year,
                    invoice_month=invoice_month,
                    closing_date=closing_date.isoformat(),
                    due_date=due_date.isoformat(),
                )
            else:
                invoice_id = invoice["id"]

            self.expense_repository.criar_lancamento(
                credit_card_id=credit_card["id"],
                invoice_id=invoice_id,
                category_id=category_id,
                effective_description=effective_description,
                effective_purchase_date=effective_purchase_date,
                billing_date=closing_date.isoformat(),
                installment_number=installment_number,
                installment_total=installment_total,
                effective_amount_cents=effective_amount_cents,
                installment_group_id=row.get("installment_group_id"),
                notes=row.get("notes"),
                original_description=original_description,
                original_purchase_date=original_purchase_date,
                original_amount_cents=original_amount_cents,
                source_type=row.get("source_type") or "portable_import",
                source_reference=row.get("source_reference"),
            )

            ja_importados_agora[assinatura] += 1
            total_importado += 1

        return total_importado